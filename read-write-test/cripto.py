from openpyxl import load_workbook
import xlrd
import xlwt

# ######txt test######

# #path = input("path:?")
# path = "C:\\Users\\Pedro\\Documents\\GitHub\\crypto_1\\test.txt"

# with open(path, "w") as file:
#     file.write("Hello, World!\n")

# ######xlsx test#####

# #path = input("path:?")
# path = "C:\\Users\\Pedro\\Documents\\GitHub\\crypto_1\\test.xlsx"

# wb = load_workbook(filename='test.xls')
# ws = wb.active

# # Iterate over rows and count non-empties
# # for row in ws.iter_rows(values_only=True):
# #     if any(cell_value is not None for cell_value in row):
# #         non_empty_row_count += 1

# row_data = []

# for row in ws.iter_rows(values_only=True):
#     if any(cell_value is not None for cell_value in row):
#         row_data.append(row)

# for row in row_data:
#         print(row)

######xls test1#####
        
# #path = input("path:?")
# path = "C:\\Users\\Pedro\\Documents\\GitHub\\crypto_1\\test.xls"
# workbook = xlrd.open_workbook(path)

# sheet = workbook.sheet_by_index(0)

# non_empty_rows = []

# for row_index in range(sheet.nrows):
#     row_data = sheet.row_values(row_index)
    
#     # Check if the row is empty
#     if any(cell.strip() for cell in row_data):
#         non_empty_rows.append(row_data)

# new_workbook = xlwt.Workbook()
# new_sheet = new_workbook.add_sheet('Sheet1')

# # Copy existing contents to the new workbook
# for sheet_index in range(workbook.nsheets):
#     existing_sheet = workbook.sheet_by_index(sheet_index)
#     for row_index in range(existing_sheet.nrows):
#         for col_index in range(existing_sheet.ncols):
#             new_sheet.write(row_index, col_index, existing_sheet.cell_value(row_index, col_index))

# # Write modified rows to the new workbook
# row_index_offset = sheet.nrows
# for row_index, row_data in enumerate(non_empty_rows ):
#     modified_row = row_data + ["?"]
#     for col_index, cell_value in enumerate(modified_row):
#         new_sheet.write(row_index + row_index_offset, col_index, cell_value)

# new_workbook.save(path)

######xls test2#####

def reverse_content(content):
    # Function to reverse the content of a cell
    return content[::-1]

def reverse_excel_file(input_file_path, output_file_path):
    # Open the original Excel file
    workbook = xlrd.open_workbook(input_file_path)
    
    # Create a new workbook to write the reversed content
    new_workbook = xlwt.Workbook()
    new_sheet = new_workbook.add_sheet('Sheet1')
    
    # Iterate over each sheet in the original workbook
    for sheet_index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_index)
        
        # Iterate over each row in the sheet
        for row_index in range(sheet.nrows):
            # Iterate over each cell in the row
            for col_index in range(sheet.ncols):
                # Read the content of the cell
                cell_content = sheet.cell_value(row_index, col_index)
                
                # Reverse the content
                reversed_content = reverse_content(cell_content)
                
                # Write the reversed content to the new workbook
                new_sheet.write(row_index, col_index, reversed_content)
    
    # Save the new workbook
    new_workbook.save(output_file_path)

# Example usage
input_file_path = "C:\\Users\\Pedro\\Documents\\GitHub\\crypto_1\\test.xls"
output_file_path = "C:\\Users\\Pedro\\Documents\\GitHub\\crypto_1\\test.xls"

reverse_excel_file(input_file_path, output_file_path)

print("Content reversed and saved to", output_file_path)


